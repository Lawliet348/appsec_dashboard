from flask import Flask, render_template, request, url_for, redirect, flash
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os

app = Flask(__name__)
EXCEL_FILE = "inputs.xlsx"

if not os.path.exists(EXCEL_FILE):
	workbook = Workbook()
	sheet = workbook.active

	# headers = ["Application name", "Criticality", "Public facing", "COTS", "Total vulnerabilities", "Vulns unresolved", "Vulns unresolved excl 'Low'", "Last tested date"]
	headers = ["Application name", "Criticality", "Public facing"]
	sheet.append(headers)

	for col_index, header in enumerate(headers, start=1):
		col_letter = sheet.cell(row=1, column=col_index).column_letter
		sheet[col_letter + "1"].font = Font(bold=True)
		sheet.column_dimensions[col_letter].width = 35

	workbook.save(EXCEL_FILE)

@app.route("/", methods=["GET", "POST"])
def dashboard():
	if request.method == "POST":
		application_name = request.form["application_name"]
		criticality = request.form["criticality"]
		public_facing = request.form["public_facing"]
		# cots = request.form["cots"]
		# num_vulns = int(request.form["num_vulns"])
		# vulns_unresolved = int(request.form["vulns_unresolved"])
		# vulns_unre_excl_low = int(request.form["vulns_unre_excl_low"])
		# last_test_date = request.form["last_test_date"]

		# if not vulns_unre_excl_low or not vulns_unresolved or not application_name or not criticality or not public_facing or not cots or not num_vulns or not last_test_date:
		if not application_name or not criticality or not public_facing:
			flash("All fields mandatory!")
			return redirect(url_for("dashboard"))

		# today = datetime.now().date()
		# last_test_date_obj = datetime.strptime(last_test_date, "%Y-%m-%d").date()
		# if last_test_date_obj > today:
		# 	flash("Cannot be a future date!")
		# 	return redirect(url_for("dashboard"))

		workbook = load_workbook(EXCEL_FILE)
		sheet = workbook.active
		# sheet.append([application_name, criticality, public_facing, cots, num_vulns, vulns_unresolved, vulns_unre_excl_low, last_test_date])
		sheet.append([application_name, criticality, public_facing])
		workbook.save(EXCEL_FILE)

		return redirect(url_for("dashboard"))
	today_date = datetime.now().strftime("%Y-%m-%d")
	return render_template("dashboard.html", today_date=today_date)

if __name__ == "__main__":
	app.run(debug=True)